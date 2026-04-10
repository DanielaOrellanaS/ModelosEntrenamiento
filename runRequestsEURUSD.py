"""
run_requests.py
Lee el log_url_AUDUSD_*.csv mas reciente en TestRequests,
lo convierte a xlsx, hace las peticiones y guarda Resultados_Request.

Uso: python run_requests.py
"""

import pandas as pd
import requests
import re
import glob
import os
from urllib.parse import urlparse, parse_qs, urlencode, urlunparse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
TEST_DIR   = os.path.join(BASE_DIR, "TestRequests")
# LOCAL_BASE = "http://10.19.15.69:80" 
LOCAL_BASE = "http://192.168.100.73:80"
SYMBOL     = "EURUSD"

TIMEOUT                  = 5    # segundos por petición
MAX_ERRORES_CONSECUTIVOS = 3    # para inmediatamente si hay N errores seguidos

# ── Buscar CSV más reciente ───────────────────────────────────
pattern = os.path.join(TEST_DIR, f"log_url_{SYMBOL}_*.csv")
files   = glob.glob(pattern)

if not files:
    raise FileNotFoundError(f"No se encontró log_url_{SYMBOL}_*.csv en:\n{TEST_DIR}")

source_csv = max(files, key=os.path.getmtime)
print(f"Leyendo: {os.path.basename(source_csv)}")

# ── Convertir CSV → XLSX ──────────────────────────────────────
df = pd.read_csv(source_csv, dtype=str, sep=None, engine="python")
df.columns = df.columns.str.strip()

xlsx_path = source_csv.replace(".csv", ".xlsx")
if not os.path.exists(xlsx_path):
    df.to_excel(xlsx_path, index=False)
    print(f"Convertido a xlsx: {os.path.basename(xlsx_path)}")
else:
    print(f"xlsx ya existe, usando: {os.path.basename(xlsx_path)}")
    df = pd.read_excel(xlsx_path, dtype=str)
    df.columns = df.columns.str.strip()

if "post" not in df.columns:
    raise ValueError(f"No se encontró la columna 'post'. Columnas: {list(df.columns)}")

print(f"Filas a procesar: {len(df)}")
print(f"Enviando a: {LOCAL_BASE}\n")

# ── Verificar que la API está corriendo ANTES de procesar ─────
print("Verificando conexión con la API...")
try:
    test = requests.get(f"{LOCAL_BASE}/health", timeout=3)
    test.raise_for_status()
    health = test.json()
    modelos = list(health.get("modelos", {}).keys())
    print(f"✓ API disponible — modelos: {modelos}\n")
except Exception as e:
    print(f"\n✗ ERROR: No se puede conectar a la API en {LOCAL_BASE}")
    print(f"  Detalle: {e}")
    print(f"  Asegúrate de que uvicorn esté corriendo antes de ejecutar este script.")
    exit(1)

# ── Adaptar URL al localhost ───────────────────────────────────
def adapt_url(original_url: str) -> str:
    url = re.sub(r'https?://[^/]+', LOCAL_BASE, original_url.strip())
    url = url.replace("?&", "?")
    parsed       = urlparse(url)
    params       = parse_qs(parsed.query, keep_blank_values=True)
    params.pop("c5d", None)
    clean_params = {k: v[0] for k, v in params.items()}
    return urlunparse(parsed._replace(query=urlencode(clean_params)))

# ── Procesar filas ────────────────────────────────────────────
results = []
errores_consecutivos = 0

for idx, row in df.iterrows():
    original_url = str(row.get("post", "")).strip()

    empty_result = {"RESULTADO":"","valor_profit":"","percentil_inf":"","percentil_sup":"",
                    "error":"","url_enviada":""}

    if not original_url or original_url.lower() == "nan":
        results.append({**empty_result, "error": "URL vacía"})
        continue

    clean_url = adapt_url(original_url)

    try:
        resp = requests.get(clean_url, timeout=TIMEOUT)
        resp.raise_for_status()
        data = resp.json()
        results.append({
            "RESULTADO":    data.get("RESULTADO", ""),
            "valor_profit": data.get("valor_profit", ""),
            "percentil_inf":data.get("percentil_inf", ""),
            "percentil_sup":data.get("percentil_sup", ""),
            "error":        "",
            "url_enviada":  clean_url,
        })
        errores_consecutivos = 0

    except requests.exceptions.ConnectionError:
        error_msg = "API no disponible — ¿está corriendo uvicorn?"
        results.append({**empty_result, "error": error_msg, "url_enviada": clean_url})
        errores_consecutivos += 1
        print(f"  ✗ Fila {idx+1}: {error_msg}")

    except Exception as e:
        error_msg = str(e)
        results.append({**empty_result, "error": error_msg, "url_enviada": clean_url})
        errores_consecutivos += 1
        print(f"  ✗ Fila {idx+1}: {error_msg}")

    # Parar si hay demasiados errores consecutivos
    if errores_consecutivos >= MAX_ERRORES_CONSECUTIVOS:
        print(f"\n⚠️  {MAX_ERRORES_CONSECUTIVOS} errores consecutivos detectados.")
        print(f"   URL fallida: {clean_url}")
        print(f"   Abortando — revisa la API y vuelve a intentar.")
        break

    if (idx + 1) % 20 == 0:
        oks = sum(1 for r in results if not r.get("error"))
        print(f"  {idx+1}/{len(df)} procesadas — OK: {oks} | Errores: {errores_consecutivos}")

print(f"\n  {len(results)}/{len(df)} procesadas. Generando Excel de resultados...")

# ── Combinar original + resultados ────────────────────────────
result_cols = ["RESULTADO","valor_profit","percentil_inf","percentil_sup","error","url_enviada"]
res_df      = pd.DataFrame(results, columns=result_cols)
combined    = pd.concat([df.reset_index(drop=True), res_df.reset_index(drop=True)], axis=1)

# ── Excel formateado ──────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Resultados"

H_FILL   = PatternFill("solid", start_color="1F3864")
BUY_FILL = PatternFill("solid", start_color="C6EFCE")
SEL_FILL = PatternFill("solid", start_color="FFDDC1")
NO_FILL  = PatternFill("solid", start_color="FFFACD")
ERR_FILL = PatternFill("solid", start_color="FFD0D0")
ALT_FILL = PatternFill("solid", start_color="F2F2F2")
thin     = Side(style="thin", color="CCCCCC")
border   = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = list(combined.columns)

for c, h in enumerate(headers, 1):
    cell           = ws.cell(row=1, column=c, value=h)
    cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill      = H_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = border
ws.row_dimensions[1].height = 30

result_start = len(df.columns) + 1

for r, (_, row_data) in enumerate(combined.iterrows(), 2):
    signal_val = str(row_data.get("RESULTADO", "")).upper()
    has_error  = str(row_data.get("error", "")).strip() not in ("", "nan")

    for c, val in enumerate(row_data, 1):
        cell           = ws.cell(row=r, column=c)
        cell.value     = None if str(val) == "nan" else val
        cell.font      = Font(name="Arial", size=9)
        cell.border    = border
        cell.alignment = Alignment(vertical="center")

        if c >= result_start:
            if has_error:                   cell.fill = ERR_FILL
            elif signal_val == "BUY":       cell.fill = BUY_FILL
            elif signal_val == "SELL":      cell.fill = SEL_FILL
            elif signal_val == "IGNORE":    cell.fill = NO_FILL
            elif r % 2 == 0:               cell.fill = ALT_FILL

            col_name = headers[c - 1]
            if col_name in ("valor_profit", "percentil_inf", "percentil_sup"):
                cell.number_format = '$#,##0.000000'

for c, col_name in enumerate(headers, 1):
    ltr = get_column_letter(c)
    if col_name in ("post", "url_enviada"):
        ws.column_dimensions[ltr].width = 55
    elif col_name in ("RESULTADO", "error"):
        ws.column_dimensions[ltr].width = 14
    elif col_name in ("valor_profit", "percentil_inf", "percentil_sup"):
        ws.column_dimensions[ltr].width = 16
    else:
        ws.column_dimensions[ltr].width = 18

ws.freeze_panes = "A2"

output_path = os.path.join(TEST_DIR, f"Resultados_Request_{SYMBOL}.xlsx")
wb.save(output_path)

buys   = sum(1 for r in results if r.get("RESULTADO") == "BUY")
sells  = sum(1 for r in results if r.get("RESULTADO") == "SELL")
nosig  = sum(1 for r in results if r.get("RESULTADO") == "IGNORE")
errors = sum(1 for r in results if r.get("error"))

print(f"✓ Guardado en: {output_path}")
print(f"  Total: {len(results)} | BUY: {buys} | SELL: {sells} | IGNORE: {nosig} | Errores: {errors}")