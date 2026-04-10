"""
run_requests_from_xlsx.py
"""

import pandas as pd
import requests
import re
import glob
import os
from urllib.parse import urlencode
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
TEST_DIR   = os.path.join(BASE_DIR, "DataFiles")
LOCAL_BASE = "http://192.168.100.73:80"
# LOCAL_BASE = "http://172.20.10.11:80"
SYMBOL     = "GBPAUD"

TIMEOUT                  = 5
MAX_ERRORES_CONSECUTIVOS = 3
MAX_FILAS                = 1000

# ── Buscar XLSX más reciente ──────────────────────────────────
pattern = os.path.join(TEST_DIR, f"Data_Entrenamiento_{SYMBOL}.xlsx")
files   = glob.glob(pattern)

if not files:
    raise FileNotFoundError(f"No se encontró Data_Entrenamiento_{SYMBOL}.xlsx en:\n{TEST_DIR}")

source_xlsx = max(files, key=os.path.getmtime)
print(f"Leyendo: {os.path.basename(source_xlsx)}")

df = pd.read_excel(source_xlsx, dtype=str)
df.columns = df.columns.str.strip()
if MAX_FILAS is not None:
    df = df.head(MAX_FILAS)
print(f"Filas a procesar: {len(df)}")

# ── Construir URL desde cada fila ─────────────────────────────
def build_url(row: pd.Series) -> str:
    """
    Construye la URL de predicción a partir de las columnas del XLSX.
    La fecha en el XLSX viene como '2025,12,03 16:20' → se convierte a '2025-12-03T16:20:00'
    """
    raw_fecha = str(row.get("fecha", "")).strip()
    # Formato fuente: "2025,12,03 16:20"  →  "2025-12-03T16:20:00"
    fecha_iso = re.sub(r"(\d{4}),(\d{2}),(\d{2})\s+(\d{2}:\d{2})", r"\1-\2-\3T\4:00", raw_fecha)

    def fmt(col):
        v = str(row.get(col, "0")).strip()
        try:
            return f"{float(v):.8f}"
        except ValueError:
            return "0.00000000"

    params = {
        "symbol": str(row.get("simbolo", SYMBOL)).strip(),
        "fecha":  fecha_iso,
        # ── Vela 5m ──────────────────────────────────────────
        "o5":  fmt("precioopen5"),
        "c5":  fmt("precioclose5"),
        "c5d": fmt("precioclose5"),   # mismo que c5 (cierre actual)
        "h5":  fmt("preciohigh5"),
        "l5":  fmt("preciolow5"),
        "v5":  fmt("volume5"),
        # ── RSI + Stochastic ─────────────────────────────────
        "r5":  fmt("rsi5"),
        "m5":  fmt("iStochaMain5"),
        "s5":  fmt("iStochaSign5"),
        # ── EMA 5m — nombres exactos del scaler ──────────────
        "ema550":          fmt("ema550"),
        "ema5200":         fmt("ema5200"),
        "ema50_prev":      fmt("ema50_prev"),
        "ema5200_prev":    fmt("ema5200_prev"),
        # ── MACD 5m — nombres exactos del scaler ─────────────
        "macdLine5":       fmt("macdLine5"),
        "signalLine5":     fmt("signalLine5"),
        "macdLine_prev5":  fmt("macdLine_prev5"),
        "signalLine_prev5":fmt("signalLine_prev5"),
        # ── ADX 5m — nombres exactos del scaler ──────────────
        "adx5":    fmt("adx5"),
        "diPlus5": fmt("diPlus5"),
        "diMinus5":fmt("diMinus5"),
        # ── 15m opcionales (aceptados pero ignorados por API) ─
        "o15": fmt("precioopen15"),
        "c15": fmt("precioclose15"),
        "h15": fmt("preciohigh15"),
        "l15": fmt("preciolow15"),
        "v15": fmt("volume15"),
        "r15": fmt("rsi15"),
        "m15": fmt("iStochaMain15"),
        "s15": fmt("iStochaSign15"),
    }

    return f"{LOCAL_BASE}/predict?{urlencode(params)}"

# ── Verificar API ─────────────────────────────────────────────
print(f"\nEnviando a: {LOCAL_BASE}")
print("Verificando conexión con la API...")
try:
    test = requests.get(f"{LOCAL_BASE}/health", timeout=3)
    test.raise_for_status()
    health = test.json()
    modelos = list(health.get("modelos", {}).keys())
    print(f"✓ API disponible — modelos: {modelos}\n")
except Exception as e:
    print(f"\n✗ ERROR: No se puede conectar a la API en {LOCAL_BASE}")
    print(f"  Detalle: {e}")
    print(f"  Asegúrate de que uvicorn esté corriendo antes de ejecutar este script.")
    exit(1)

# ── Procesar filas ────────────────────────────────────────────
results = []
errores_consecutivos = 0
empty_result = {"RESULTADO":"","valor_profit":"","percentil_inf":"","percentil_sup":"",
                "error":"","url_enviada":""}

for idx, row in df.iterrows():
    clean_url = build_url(row)

    try:
        resp = requests.get(clean_url, timeout=TIMEOUT)
        resp.raise_for_status()
        data = resp.json()
        results.append({
            "RESULTADO":    data.get("RESULTADO", ""),
            "valor_profit": data.get("valor_profit", ""),
            "percentil_inf":data.get("percentil_inf", ""),
            "percentil_sup":data.get("percentil_sup", ""),
            "error":        "",
            "url_enviada":  clean_url,
        })
        errores_consecutivos = 0

    except requests.exceptions.ConnectionError:
        error_msg = "API no disponible — ¿está corriendo uvicorn?"
        results.append({**empty_result, "error": error_msg, "url_enviada": clean_url})
        errores_consecutivos += 1
        print(f"  ✗ Fila {idx+1}: {error_msg}")

    except Exception as e:
        error_msg = str(e)
        results.append({**empty_result, "error": error_msg, "url_enviada": clean_url})
        errores_consecutivos += 1
        print(f"  ✗ Fila {idx+1}: {error_msg}")

    if errores_consecutivos >= MAX_ERRORES_CONSECUTIVOS:
        print(f"\n⚠️  {MAX_ERRORES_CONSECUTIVOS} errores consecutivos detectados.")
        print(f"   URL fallida: {clean_url}")
        print(f"   Abortando — revisa la API y vuelve a intentar.")
        break

    if (idx + 1) % 100 == 0:
        oks = sum(1 for r in results if not r.get("error"))
        print(f"  {idx+1}/{len(df)} procesadas — OK: {oks} | Errores: {errores_consecutivos}")

print(f"\n  {len(results)}/{len(df)} procesadas. Generando Excel de resultados...")

# ── Combinar original + resultados ────────────────────────────
result_cols = ["RESULTADO","valor_profit","percentil_inf","percentil_sup","error","url_enviada"]
res_df   = pd.DataFrame(results, columns=result_cols)
combined = pd.concat([df.reset_index(drop=True), res_df.reset_index(drop=True)], axis=1)

# ── Excel formateado ──────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Resultados"

H_FILL   = PatternFill("solid", start_color="1F3864")
BUY_FILL = PatternFill("solid", start_color="C6EFCE")
SEL_FILL = PatternFill("solid", start_color="FFDDC1")
NO_FILL  = PatternFill("solid", start_color="FFFACD")
ERR_FILL = PatternFill("solid", start_color="FFD0D0")
ALT_FILL = PatternFill("solid", start_color="F2F2F2")
thin     = Side(style="thin", color="CCCCCC")
border   = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = list(combined.columns)

for c, h in enumerate(headers, 1):
    cell           = ws.cell(row=1, column=c, value=h)
    cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill      = H_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = border
ws.row_dimensions[1].height = 30

result_start = len(df.columns) + 1

for r, (_, row_data) in enumerate(combined.iterrows(), 2):
    signal_val = str(row_data.get("RESULTADO", "")).upper()
    has_error  = str(row_data.get("error", "")).strip() not in ("", "nan")

    for c, val in enumerate(row_data, 1):
        cell           = ws.cell(row=r, column=c)
        cell.value     = None if str(val) == "nan" else val
        cell.font      = Font(name="Arial", size=9)
        cell.border    = border
        cell.alignment = Alignment(vertical="center")

        if c >= result_start:
            if has_error:                cell.fill = ERR_FILL
            elif signal_val == "BUY":    cell.fill = BUY_FILL
            elif signal_val == "SELL":   cell.fill = SEL_FILL
            elif signal_val == "IGNORE": cell.fill = NO_FILL
            elif r % 2 == 0:            cell.fill = ALT_FILL

            col_name = headers[c - 1]
            if col_name in ("valor_profit", "percentil_inf", "percentil_sup"):
                cell.number_format = '$#,##0.000000'

for c, col_name in enumerate(headers, 1):
    ltr = get_column_letter(c)
    if col_name == "url_enviada":
        ws.column_dimensions[ltr].width = 55
    elif col_name in ("RESULTADO", "error"):
        ws.column_dimensions[ltr].width = 14
    elif col_name in ("valor_profit", "percentil_inf", "percentil_sup"):
        ws.column_dimensions[ltr].width = 16
    elif col_name == "fecha":
        ws.column_dimensions[ltr].width = 20
    else:
        ws.column_dimensions[ltr].width = 14

ws.freeze_panes = "A2"

os.makedirs(TEST_DIR, exist_ok=True)
output_path = os.path.join(TEST_DIR, f"Resultados_Request_{SYMBOL}.xlsx")
wb.save(output_path)

buys   = sum(1 for r in results if r.get("RESULTADO") == "BUY")
sells  = sum(1 for r in results if r.get("RESULTADO") == "SELL")
nosig  = sum(1 for r in results if r.get("RESULTADO") == "IGNORE")
errors = sum(1 for r in results if r.get("error"))

print(f"✓ Guardado en: {output_path}")
print(f"  Total: {len(results)} | BUY: {buys} | SELL: {sells} | IGNORE: {nosig} | Errores: {errors}")